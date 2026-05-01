# Day 25 - Pivot-style analysis using only openpyxl
# NO pandas — that's Day 29. Only tools from Phase 1.

from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Alignment

# ─────────────────────────────────────────────
# STEP 1: Load the practice file and read all rows
# (You learned this on Day 7)
# ─────────────────────────────────────────────
wb = load_workbook("Day25_PivotTables_Practice.xlsx")
ws = wb["Sales_Data"]

headers = [cell.value for cell in ws[1]]  # Get column names from row 1
rows = []
for row in ws.iter_rows(min_row=2, values_only=True):
    if row[0] is None:
        continue
    rows.append(dict(zip(headers, row)))  # Each row → dict like {Region: "North", Revenue: 60000}

# ─────────────────────────────────────────────
# STEP 2: Group & Sum by Region (manual pivot)
# This is what Excel Pivot Table does internally
# ─────────────────────────────────────────────
region_revenue = {}
for r in rows:
    region = r["Region"]
    revenue = r["Revenue"] or 0
    region_revenue[region] = region_revenue.get(region, 0) + revenue

# ─────────────────────────────────────────────
# STEP 3: Group & Sum by Salesperson
# ─────────────────────────────────────────────
person_revenue = {}
for r in rows:
    person = r["Salesperson"]
    person_revenue[person] = person_revenue.get(person, 0) + (r["Revenue"] or 0)

# ─────────────────────────────────────────────
# STEP 4: Group & AVERAGE Profit Margin by Product
# Value Field Settings → AVERAGE in Excel
# ─────────────────────────────────────────────
product_margins = {}
product_counts  = {}
for r in rows:
    product = r["Product"]
    margin  = r["Profit_Margin%"] or 0
    product_margins[product] = product_margins.get(product, 0) + margin
    product_counts[product]  = product_counts.get(product, 0) + 1

# ─────────────────────────────────────────────
# STEP 5: 2D Grouping — Region × Quarter
# Like dragging both Region to Rows and Quarter to Columns in Excel
# ─────────────────────────────────────────────
pivot_2d = {}
quarters = set()
for r in rows:
    region  = r["Region"]
    quarter = r["Quarter"]
    revenue = r["Revenue"] or 0
    quarters.add(quarter)
    if region not in pivot_2d:
        pivot_2d[region] = {}
    pivot_2d[region][quarter] = pivot_2d[region].get(quarter, 0) + revenue

quarters = sorted(quarters)

# ─────────────────────────────────────────────
# STEP 6: Write all summaries to a new Excel file
# (You learned this on Day 8)
# ─────────────────────────────────────────────
wb_out = Workbook()
ws_out = wb_out.active
ws_out.title = "Pivot_Summary"

ws_out["A1"] = "Day 25 - Pivot Analysis Output"
ws_out["A1"].font = Font(bold=True, size=13, color="1F4E79")

# Region Summary
row = 3
ws_out.cell(row=row, column=1, value="Region").font = Font(bold=True)
ws_out.cell(row=row, column=2, value="Revenue").font = Font(bold=True)
row += 1
for region, total in sorted(region_revenue.items(), key=lambda x: -x[1]):
    ws_out.cell(row=row, column=1, value=region)
    ws_out.cell(row=row, column=2, value=total)
    ws_out.cell(row=row, column=2).number_format = '₹#,##0'
    row += 1

# Salesperson Summary
row += 2
ws_out.cell(row=row, column=1, value="Salesperson").font = Font(bold=True)
ws_out.cell(row=row, column=2, value="Revenue").font = Font(bold=True)
row += 1
for person, total in sorted(person_revenue.items(), key=lambda x: -x[1]):
    ws_out.cell(row=row, column=1, value=person)
    ws_out.cell(row=row, column=2, value=total)
    ws_out.cell(row=row, column=2).number_format = '₹#,##0'
    row += 1

# Product Margin Summary
row += 2
ws_out.cell(row=row, column=1, value="Product").font = Font(bold=True)
ws_out.cell(row=row, column=2, value="Avg Profit Margin%").font = Font(bold=True)
row += 1
for product in sorted(product_margins, key=lambda x: -product_margins[x]/product_counts[x]):
    avg = round(product_margins[product] / product_counts[product], 1)
    ws_out.cell(row=row, column=1, value=product)
    ws_out.cell(row=row, column=2, value=f"{avg}%")
    row += 1

# 2D Pivot Summary - Region × Quarter
row += 2
ws_out.cell(row=row, column=1, value="Region").font = Font(bold=True)
col = 2
for q in quarters:
    ws_out.cell(row=row, column=col, value=q).font = Font(bold=True)
    col += 1
row += 1
for region in sorted(pivot_2d):
    ws_out.cell(row=row, column=1, value=region)
    col = 2
    for q in quarters:
        val = pivot_2d[region].get(q, 0)
        ws_out.cell(row=row, column=col, value=val)
        ws_out.cell(row=row, column=col).number_format = '₹#,##0'
        col += 1
    row += 1

wb_out.save("Day25_Pivot_Output.xlsx")